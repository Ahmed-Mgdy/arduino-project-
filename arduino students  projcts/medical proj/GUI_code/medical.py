import serial
import serial.tools.list_ports
import threading
import tkinter as tk    
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import subprocess

EXCEL_FILE = "sensor_data.xlsx"

class SensorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sensor Monitor with Excel Logging")
        self.ser = None
        self.running = False
        self.latest_data = {}
        self.all_data = []  # Store all sensor data

        self.create_widgets()
        self.setup_excel()
        self.detect_ports()

    def create_widgets(self):
        self.root.configure(bg="#2f4f4f")  # Background color for root

        # COM Port selection
        port_frame = tk.Frame(self.root, bg="#2f4f4f")
        port_frame.pack(pady=20)

        tk.Label(port_frame, text="Select COM Port:", font=("Helvetica", 12, "bold"), fg="#ffffff", bg="#2f4f4f").pack(side=tk.LEFT)
        self.port_combo = ttk.Combobox(port_frame, width=20, state="readonly", font=("Helvetica", 12))
        self.port_combo.pack(side=tk.LEFT, padx=10)
        self.connect_button = self.create_button(port_frame, "Connect", self.connect_serial)
        self.connect_button.pack(side=tk.LEFT, padx=10)

        # Labels for sensor values
        self.ambient_label = self.create_label(self.root, "Ambient Temp: -- °C")
        self.object_label = self.create_label(self.root, "Object Temp: -- °C")
        self.hr_label = self.create_label(self.root, "Heart Rate: -- bpm")
        self.spo2_label = self.create_label(self.root, "SpO2: -- %")
        self.beat_label = tk.Label(self.root, text="", font=("Arial", 16), fg="red", bg="#2f4f4f")
        self.beat_label.pack(pady=20)

        # Log box
        self.log_text = tk.Text(self.root, height=10, width=80, font=("Courier", 12), bg="#f0f0f0", fg="#000000", bd=2, relief="groove")
        self.log_text.pack(pady=20)

        # Control buttons
        button_frame = tk.Frame(self.root, bg="#2f4f4f")
        button_frame.pack(pady=10)

        self.save_button = self.create_button(button_frame, "Save to Excel", self.save_to_excel)
        self.save_button.pack(side=tk.LEFT, padx=10)

        self.open_button = self.create_button(button_frame, "Open Excel File", self.open_excel)
        self.open_button.pack(side=tk.LEFT, padx=10)

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def create_button(self, parent, text, command):
        button = tk.Button(parent, text=text, command=command, font=("Helvetica", 12), fg="white", bg="#4682b4", relief="flat", width=15, height=2)
        button.bind("<Enter>", self.on_hover)
        button.bind("<Leave>", self.on_leave)
        return button

    def on_hover(self, event):
        event.widget.config(bg="#5a9bd4")  # Hover effect

    def on_leave(self, event):
        event.widget.config(bg="#4682b4")  # Reset color

    def create_label(self, parent, text):
        label = tk.Label(parent, text=text, font=("Helvetica", 14), fg="#ffffff", bg="#2f4f4f")
        label.pack(pady=5)
        return label

    def detect_ports(self):
        ports = serial.tools.list_ports.comports()
        port_list = [port.device for port in ports]
        self.port_combo['values'] = port_list
        if port_list:
            self.port_combo.current(0)

    def setup_excel(self):
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            wb.save(EXCEL_FILE)

    def connect_serial(self):
        selected_port = self.port_combo.get()
        if not selected_port:
            messagebox.showwarning("Warning", "Please select a COM port.")
            return
        try:
            self.ser = serial.Serial(selected_port, 115200, timeout=1)
            self.running = True
            self.read_thread = threading.Thread(target=self.read_serial)
            self.read_thread.daemon = True
            self.read_thread.start()
            self.connect_button.config(state="disabled")
            self.port_combo.config(state="disabled")
        except Exception as e:
            messagebox.showerror("Connection Error", str(e))

    def read_serial(self):
        while self.running:
            try:
                line = self.ser.readline().decode().strip()
                if not line:
                    continue

                if "Beat" in line:
                    self.beat_label.config(text="♥ Beat!")
                    self.root.after(500, lambda: self.beat_label.config(text=""))
                    self.log_text.insert(tk.END, "♥ Beat!\n")
                    self.log_text.see(tk.END)
                    continue

                if "Ambient" in line:
                    parts = line.replace("°C", "").replace("bpm", "").replace("%", "").split("|")
                    data = {}
                    for part in parts:
                        if ":" in part:
                            key, value = part.strip().split(":")
                            data[key.strip()] = float(value.strip())

                    self.latest_data = data
                    self.all_data.append(data)  # Add the data to the list
                    self.update_gui(data)

                    log_line = f"{datetime.now().strftime('%H:%M:%S')} -> {line}"
                    self.log_text.insert(tk.END, log_line + "\n")
                    self.log_text.see(tk.END)

            except Exception as e:
                print(f"Serial Read Error: {e}")

    def update_gui(self, data):
        self.ambient_label.config(text=f"Ambient Temp: {data.get('Ambient', '--'):.2f} °C")
        self.object_label.config(text=f"Object Temp: {data.get('Object', '--'):.2f} °C")
        self.hr_label.config(text=f"Heart Rate: {data.get('HR', '--'):.2f} bpm")
        self.spo2_label.config(text=f"SpO2: {data.get('SpO2', '--'):.2f} %")

    def save_to_excel(self):
        if not self.all_data:
            messagebox.showinfo("No Data", "No sensor data available yet.")
            return
        try:
            # Load the existing workbook
            wb = load_workbook(EXCEL_FILE)
            
            # Create a new sheet with a unique name (current date and time)
            sheet_name = datetime.now().strftime("Data_%Y-%m-%d_%H-%M-%S")
            
            # Add a new sheet to the workbook
            ws = wb.create_sheet(sheet_name)
            
            # Add headers to the new sheet
            ws.append(["Time", "Ambient Temp (°C)", "Object Temp (°C)", "Heart Rate (bpm)", "SpO2 (%)"])

            # Add the data to the new sheet
            for data in self.all_data:
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([
                    now,
                    data.get('Ambient', 0),
                    data.get('Object', 0),
                    data.get('HR', 0),
                    data.get('SpO2', 0)
                ])
            
            # Save the workbook
            wb.save(EXCEL_FILE)
            
            # Clear all_data after saving to ensure it's ready for the next session
            self.all_data = []
            
            messagebox.showinfo("Saved", "All sensor data saved to a new sheet in Excel.")
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    def open_excel(self):
        try:
            if os.name == 'nt':
                os.startfile(EXCEL_FILE)
            elif os.name == 'posix':
                subprocess.call(['open', EXCEL_FILE])
            else:
                subprocess.call(['xdg-open', EXCEL_FILE])
        except Exception as e:
            messagebox.showerror("Open File Error", str(e))

    def on_closing(self):
        self.running = False
        if self.ser and self.ser.is_open:
            self.ser.close()
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = SensorApp(root)
    root.mainloop()
